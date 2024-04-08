# -*- coding: utf-8 -*-
"""
Created on Sat Mar  9 14:49:51 2024

@author: Kun Zhang, Drew Mortenson
"""

### This code is used to calculate temperature field of asphalt pavement ###

import openpyxl as xl
import math
import numpy
import shared
import matplotlib.pyplot as plt
#from scipy import stats
from sklearn.linear_model import LinearRegression
import pandas as pd
from datetime import datetime

plt.switch_backend('agg')

row_correction = 0  # shift peak of test to match the peak of simulation, default = 0

# delt_e = e_a - e_AC, difference between absorption and emission of pavement surface, negative value, seasonal adjustment
delta_e_1 = 0.0000000000000 #-0.15  # winter time value in Jan and Dec # default = -0.15
delta_e_6 = 0.0000000000000 #-0.12  # summer time value in June and July # default = -0.05

# constants used in the simulation, e.g. density (rho) of pavement
rho_AC = 2450  # density kg/m3; value from Omairey et al. 2021, this can be found from LTPP
rho_base = 2350
rho_subbase = 2350
rho_subgrade = 2200
rho_water = 1000

k_AC = 2.0  # thermal conductivity, W/m/K; range 1-2.88, Mid = 1.94 (Omairey et al. 2022)
k_base = 1.75  # range = 1-2.5, Mid = 1.75
k_subbase = 1.75 # Ditto
k_subgrade = 1.345 # range = 0.19-2.5, Mid = 1.345
k_water = 0.607

cp_AC = 950  # heat capacity, J/kg/K range = 700-1200; Mid = 950 (Omairey et al. 2022)
cp_base = 900 # Range = 700 - 1100, Mid = 900
cp_subbase = 900 # Ditto
cp_subgrade = 850 # range = 600-1100, Mid = 850
cp_water = 4182

alpha_AC = k_AC / rho_AC / cp_AC * 3600  # heat diffusivity  convert m2/s to m2/h
alpha_base = k_base / rho_base / cp_base * 3600
alpha_subbase = k_subbase / rho_subbase / cp_subbase * 3600
alpha_subgrade = k_subgrade / rho_subgrade / cp_subgrade * 3600
alpha_water = k_water / rho_water / cp_water * 3600

def run_simulation(post_process, Ucode, solarFile=None, windFile=None, tempFile=None, Thermo_depth=[],
                   thickness_AC=0, thickness_Base=0, thickness_subbase=0, thickness_subgrade=0,
                   delta_e_1=-.15, delta_e_6=-.12):
    # find environmental data, e.g T_air, from excel

    env_temp_data = xl.load_workbook(tempFile)
    temp_sheet = env_temp_data['MERRA_TEMP_HOUR']

    env_wind_data = xl.load_workbook(windFile)
    wind_sheet = env_wind_data['MERRA_WIND_HOUR']

    env_solar_data = xl.load_workbook(solarFile)
    solar_sheet = env_solar_data['MERRA_SOLAR_HOUR']


    # Define model constant
    delta_z = 0.001  # each sub_layer is 1 mm for AC, Base, and Subbase
    delta_z_subgrade = 0.001  # sub_layer is 1 mm for subgrade; could use 5 mm

    N_AC = int(thickness_AC / delta_z)
    N_base = int(thickness_Base / delta_z)
    N_subbase = int(thickness_subbase / delta_z)
    N_subgrade = int(thickness_subgrade / delta_z_subgrade)
    N_total = N_AC + N_base + N_subbase + N_subgrade

    #print(N_total)

    delta_t = 1  # timestep unit = hours, need to be consistent with alpha (m2/h)
    t_total = temp_sheet.max_row - 1 #int(365 * 24)   # total simulation time (hour)
    # Define initial condition of Temperature

    T_surf = temp_sheet.cell(2, 4).value + 273.15  # Define initial surface temp is T_air, unit Kevin
    T_bott = 30 + 273.15  # Define initial bottom temperature
    initial_slope = (T_bott - T_surf) / 3
    T = numpy.zeros((N_total + 2, t_total + 1))  # Define initial Temperature_Field, T[0] = T_surf, surface Temp; T[-1] = T_bott, bottom Temp
    T[0, 0] = T_surf
    T[N_total + 1, 0] = T_bott

    for N_ele in range(1, N_total - N_subgrade + 1):  # N_ele is element number, update from AC to bottom of subbase
        T[N_ele, 0] = initial_slope * (N_ele * delta_z - delta_z / 2) + T_surf  # location of T1 is at 0.5 mm
    for N_ele in range(N_total - N_subgrade + 1, N_total + 1):  # update temperature for subgrade
        T[N_ele, 0] = initial_slope * (thickness_AC + thickness_Base + thickness_subbase + (
                    N_ele - (N_AC + N_base + N_subbase)) * delta_z_subgrade - delta_z_subgrade / 2) + T_surf

    # Define coefficient matrix from element 1 to N
    a = [0] * (N_total + 2)
    b = [0] * (N_total + 2)
    c = [0] * (N_total + 2)
    d = [0] * (N_total + 2)
    P = [0] * (N_total + 2)
    Q = [0] * (N_total + 2)

    T_star = [0] * (N_total + 2)  # Define T* as temporal T matrix for iteration of non-linear problem

    sigma = 5.67e-8  # Stefen-Boltzmann constant W/m2/K4

    for time in range(1, t_total + 1):  # need to run "t_total+1"  ; test case for 24 hours
        # End Early Condition
        if (shared.endEarly.is_set()):
            return 0
            # sys.exit()   
    
        i_flag = 1
        i_iter = 1

        for N_ele in range(0, N_total + 2):
            T_star[N_ele] = T[N_ele, time - 1]  # assign previous step values to T_star matrix (temporal matrix)

        #print('T_surf:', T_star[0])
        #print('T_bott', T_star[N_total + 1])

        e_AC = 0.85 # emissivity of pavement;
        # delt_e = e_a - e_AC, difference between absorption and emission, negative value, seasonal
        # delta_e_1 = -0.20  # winter time value in Jan and Dec # default = -0.15
        # delta_e_6 = -0.05  # summer time value in June and July # default = -0.05
        A = numpy.array([[1, 1, 1], [36, 6, 1], [144, 12, 1]])
        B = numpy.array([delta_e_1, delta_e_6, delta_e_1])
        coef = numpy.linalg.inv(A).dot(B)

        date = str(solar_sheet.cell(time + 1, 2).value)
        month = int(date[5])*10 + int(date[6])

        print(date)

        delta_e = coef[0] * month**2 + coef[1] * month + coef[2]  # seasonal e_a as parabolic curve

        albedo_AC = 0.2  # vary from 0.15-0.35; see more in Han et al. 2011;
        # Values are from MERRA
        # if solar_sheet.cell(time + 1, 8).value is None:
        #    albedo_AC = 0
        # if solar_sheet.cell(time + 1, 8).value is not None:
        #    albedo_AC = solar_sheet.cell(time + 1, 8).value

        Q_solar = solar_sheet.cell(time + 1, 4).value  # get short_wave surface (W/m2);
        q_solar = (1 - albedo_AC) * Q_solar
        T_air = temp_sheet.cell(time + 1, 4).value + 273.15
        q_rad_diff = delta_e * sigma * (T_air ** 4)  # delta_e = (e_a - e_AC)
        v_wind = wind_sheet.cell(time + 1, 4).value  # wind velocity (m/s)

        while i_flag == 1 and i_iter <= 200:  # iteration max
            # End Early Condition
            if (shared.endEarly.is_set()):
                return 0
                # sys.exit()

            # prescript surface boundary condition

            h_conv = 698.24 * 1.4 * (0.00144 * ((abs((T_air + T_star[0]) / 2)) ** (0.3)) * (v_wind ** 0.5) + 0.00097 * (
                    abs(T_star[0] - T_air) ** 0.3))

            h_rad = sigma * e_AC * (T_star[0] ** 2 + T_air ** 2) * (T_star[0] + T_air)


            T_star[0] = (q_solar + (h_conv + h_rad) * T_air + q_rad_diff + k_AC / (delta_z / 2) * T_star[1]) / (h_conv + h_rad + k_AC / (delta_z / 2))


            #print('T0_update:', T_star[0])

            # prescript bottom boundary condition
            T_star[N_total + 1] = T_star[N_total] + delta_z_subgrade / 2 * (-2.3)

            #print('T_bott_update:', T_star[N_total + 1])

            for N_ele in range(1, N_total + 1):
                # End Early Condition
                if (shared.endEarly.is_set()):
                    return 0
                    # sys.exit()

                # initialize coefficient matrix a, b, c, d from element 1 to N; # linearization of surface BC
                if N_ele == 1:
                    b[1] = alpha_AC / delta_z
                    c[1] = 0
                    a[1] = alpha_AC / delta_z + delta_z / delta_t + 2 * alpha_AC / delta_z - (alpha_AC / (delta_z/2)) * (k_AC / (delta_z/2)) / (h_rad + h_conv + k_AC/(delta_z/2))
                    d[1] = T[1, time - 1] * delta_z / delta_t + (alpha_AC / (delta_z/2)) * (q_solar + + (h_rad+h_conv) * T_air + q_rad_diff) / (h_rad + h_conv + k_AC/(delta_z/2))

                if 1 < N_ele < N_AC:
                    b[N_ele] = alpha_AC / delta_z
                    c[N_ele] = alpha_AC / delta_z
                    a[N_ele] = (delta_z / delta_t) + 2 * alpha_AC / delta_z
                    d[N_ele] = (delta_z / delta_t) * T[N_ele, time - 1]

                if N_ele == N_AC:  # at AC and Base interface
                    #b[N_ele] = ((k_AC + k_base) / 2) / ((rho_AC * cp_AC + rho_base * cp_base) / 2) / delta_z
                    b[N_ele] = (2*alpha_AC*alpha_base) / (alpha_AC+alpha_base) / delta_z
                    c[N_ele] = alpha_AC / delta_z
                    a[N_ele] = alpha_AC / delta_z + (2*alpha_AC*alpha_base) / (alpha_AC+alpha_base) / delta_z + (delta_z / delta_t)
                    d[N_ele] = (delta_z / delta_t) * T[N_ele, time - 1]

                if N_ele == N_AC + 1:  # at AC and Base interface
                    b[N_ele] = alpha_AC / delta_z
                    c[N_ele] = (2*alpha_AC*alpha_base) / (alpha_AC+alpha_base) / delta_z
                    a[N_ele] = alpha_AC / delta_z + (2*alpha_AC*alpha_base) / (alpha_AC+alpha_base) / delta_z + (delta_z / delta_t)
                    d[N_ele] = (delta_z / delta_t) * T[N_ele, time - 1]

                if N_AC + 1 < N_ele < N_AC + N_base:
                    b[N_ele] = alpha_base / delta_z
                    c[N_ele] = alpha_base / delta_z
                    a[N_ele] = (delta_z / delta_t) + 2 * alpha_base / delta_z
                    d[N_ele] = (delta_z / delta_t) * T[N_ele, time - 1]

                if N_ele == N_AC + N_base:  # at base and subbase interface
                    b[N_ele] = (2*alpha_base+alpha_subbase) / (alpha_base + alpha_subbase) / delta_z
                    c[N_ele] = alpha_base / delta_z
                    a[N_ele] = alpha_base / delta_z + (2*alpha_base+alpha_subbase) / (alpha_base + alpha_subbase) / delta_z + (delta_z / delta_t)
                    d[N_ele] = (delta_z / delta_t) * T[N_ele, time - 1]

                if N_ele == N_AC + N_base + 1:  # at base and subbase interface
                    b[N_ele] = alpha_subbase / delta_z
                    c[N_ele] = (2*alpha_base+alpha_subbase) / (alpha_base + alpha_subbase) / delta_z
                    a[N_ele] = alpha_subbase / delta_z + (2*alpha_base+alpha_subbase) / (alpha_base + alpha_subbase) / delta_z + (delta_z / delta_t)
                    d[N_ele] = (delta_z / delta_t) * T[N_ele, time - 1]

                if N_AC + N_base + 1 < N_ele < N_AC + N_base + N_subbase:
                    b[N_ele] = alpha_subbase / delta_z
                    c[N_ele] = alpha_subbase / delta_z
                    a[N_ele] = (delta_z / delta_t) + 2 * alpha_subbase / delta_z
                    d[N_ele] = (delta_z / delta_t) * T[N_ele, time - 1]

                if N_ele == N_AC + N_base + N_subbase:  # at subbase and subgrade interface
                    b[N_ele] = (alpha_subbase*alpha_subgrade*0.5*(delta_z+delta_z_subgrade)) / (alpha_subbase*0.5*delta_z_subgrade + alpha_subgrade*0.5*delta_z) / (0.5 * (delta_z + delta_z_subgrade))
                    c[N_ele] = alpha_subbase / delta_z
                    a[N_ele] = alpha_subbase / delta_z + (alpha_subbase*alpha_subgrade*0.5*(delta_z+delta_z_subgrade)) / (alpha_subbase*0.5*delta_z_subgrade + alpha_subgrade*0.5*delta_z) / (0.5 * (delta_z + delta_z_subgrade)) + (delta_z / delta_t)
                    d[N_ele] = (delta_z / delta_t) * T[N_ele, time - 1]

                if N_ele == N_AC + N_base + N_subbase + 1:  # at subbase and subgrade interface
                    b[N_ele] = alpha_subgrade / delta_z_subgrade
                    c[N_ele] = (alpha_subbase*alpha_subgrade*0.5*(delta_z+delta_z_subgrade)) / (alpha_subbase*0.5*delta_z_subgrade + alpha_subgrade*0.5*delta_z) / (0.5 * (delta_z + delta_z_subgrade))
                    a[N_ele] = alpha_subgrade / delta_z_subgrade + (alpha_subbase*alpha_subgrade*0.5*(delta_z+delta_z_subgrade)) / (alpha_subbase*0.5*delta_z_subgrade + alpha_subgrade*0.5*delta_z) / (0.5 * (delta_z + delta_z_subgrade)) + (delta_z_subgrade / delta_t)
                    d[N_ele] = (delta_z_subgrade / delta_t) * T[N_ele, time - 1]

                if N_base == 0 and N_subbase == 0 and N_ele == N_AC:   # In case of no base and subbase
                    b[N_ele] = (alpha_AC * alpha_subgrade * 0.5 * (delta_z + delta_z_subgrade)) / (alpha_AC * 0.5 * delta_z_subgrade + alpha_subgrade * 0.5 * delta_z) / (
                                           0.5 * (delta_z + delta_z_subgrade))
                    c[N_ele] = alpha_subbase / delta_z
                    a[N_ele] = alpha_subbase / delta_z + (alpha_AC * alpha_subgrade * 0.5 * (delta_z + delta_z_subgrade)) / (
                                alpha_AC * 0.5 * delta_z_subgrade + alpha_subgrade * 0.5 * delta_z) / (
                                           0.5 * (delta_z + delta_z_subgrade)) + (delta_z / delta_t)
                    d[N_ele] = (delta_z / delta_t) * T[N_ele, time - 1]

                if N_base == 0 and N_subbase == 0 and N_ele == N_AC+1:   # In case of no base and subbase
                    b[N_ele] = alpha_subgrade / delta_z_subgrade
                    c[N_ele] = (alpha_AC*alpha_subgrade*0.5*(delta_z+delta_z_subgrade)) / (alpha_AC*0.5*delta_z_subgrade + alpha_subgrade*0.5*delta_z) / (0.5 * (delta_z + delta_z_subgrade))
                    a[N_ele] = alpha_subgrade / delta_z_subgrade + (alpha_AC*alpha_subgrade*0.5*(delta_z+delta_z_subgrade)) / (alpha_AC*0.5*delta_z_subgrade + alpha_subgrade*0.5*delta_z) / (0.5 * (delta_z + delta_z_subgrade)) + (delta_z_subgrade / delta_t)
                    d[N_ele] = (delta_z_subgrade / delta_t) * T[N_ele, time - 1]

                if N_AC + N_base + N_subbase + 1 < N_ele < N_total:
                    b[N_ele] = alpha_subgrade / delta_z_subgrade
                    c[N_ele] = alpha_subgrade / delta_z_subgrade
                    a[N_ele] = (delta_z_subgrade / delta_t) + 2 * alpha_subgrade / delta_z_subgrade
                    d[N_ele] = (delta_z_subgrade / delta_t) * T[N_ele, time - 1]

                if N_ele == N_total:
                    b[N_ele] = 0
                    c[N_ele] = alpha_subgrade / delta_z_subgrade
                    a[N_ele] = alpha_subgrade / delta_z_subgrade + delta_z_subgrade / delta_t + 2 * alpha_subgrade / delta_z_subgrade - 2 * alpha_subgrade / delta_z_subgrade
                    d[N_ele] = T[N_ele, time - 1] * delta_z_subgrade / delta_t + alpha_subgrade * (-2.3)

                # forward substitution
                P[1] = b[1] / a[1]
                Q[1] = d[1] / a[1]
            # print(P[1], Q[1], time, a[2], b[2], c[2], d[2])
            for N_ele in range(2, N_total + 1):
                P[N_ele] = b[N_ele] / (a[N_ele] - c[N_ele] * P[N_ele - 1])
                Q[N_ele] = (d[N_ele] + c[N_ele] * Q[N_ele - 1]) / (a[N_ele] - c[N_ele] * P[N_ele - 1])

            # backward substitution
            T_star[N_total] = Q[N_total]
            for N_ele in range(N_total - 1, 0, -1):
                T_star[N_ele] = P[N_ele] * T_star[N_ele + 1] + Q[N_ele]

            # check convergence
            error_T = 0
            for N_ele in range(0, N_total + 2):
                error_T += abs(T_star[N_ele] - T[N_ele, time]) ** 2
            
            # End Early Condition
            if (shared.endEarly.is_set()):
                return 0
                # sys.exit()

            #print(error_T)

            if error_T >= 0.001:
                i_iter += 1
                i_flag = 1

                for N_ele in range(0, N_total + 2):  # update temperature
                    T[N_ele, time] = T_star[N_ele]

            elif error_T < 0.001:
                i_flag = 0
                # update temperature
                for N_ele in range(0, N_total + 2):
                    T[N_ele, time] = T_star[N_ele]

            # print(time, i_iter)
            # print(T_star)
            
    # print out the results and save in excel
    results_path = shared.proj_folder_long

    results_name = "\\" + shared.proj_name + '-Simulation.xlsx'

    try:
        wb_result = xl.load_workbook(results_path + results_name)
    except FileNotFoundError:
        wb_result = xl.Workbook()  # create wb as workbook
        ws_Sheet = wb_result.active  # define ws as worksheet
        ws_Sheet.title = 'Sheet1'
        wb_result.save(results_path + results_name)
    wb_result = xl.load_workbook(results_path + results_name)
    ws_Sheet = wb_result['Sheet1']

    ws_Sheet.cell(1, 1).value = "Date"
    ws_Sheet.cell(1, 2).value = "time"
    ws_Sheet.cell(1, 3).value = 'surface'
    # ws_Sheet.cell(1, 4).value = 'Thermo_1'
    # ws_Sheet.cell(1, 5).value = 'Thermo_2'
    # ws_Sheet.cell(1, 6).value = 'Thermo_3'
    # ws_Sheet.cell(1, 7).value = 'Thermo_4'
    # ws_Sheet.cell(1, 8).value = 'Thermo_5'
    # ws_Sheet.cell(1, 9).value = 'Thermo_6'
    # ws_Sheet.cell(1, 10).value = 'Thermo_7'
    # ws_Sheet.cell(1, 11).value = 'Thermo_8'
    # ws_Sheet.cell(1, 12).value = 'Thermo_9'
    # ws_Sheet.cell(1, 13).value = 'Thermo_10'
    itr = 4
    for depth in Thermo_depth:
        ws_Sheet.cell(1, itr).value = str(depth) + ' m'
        itr += 1

    ws_Sheet.cell(2, 1).value = 0 # Intial time 0

    for i_time in range(3, t_total + 3):   # use t_total
        #ws_Sheet.cell(i_time, 1).value = i_time - 2
        ws_Sheet.cell(i_time, 1).value = solar_sheet.cell(i_time-1, 2).value
        ws_Sheet.cell(i_time, 2).value = solar_sheet.cell(i_time-1, 3).value

    for thermo_id in range (4, len(Thermo_depth)+4):  # this is column
        for row in range(2, t_total + 3):

            #ws_Sheet.cell(row, 3).value = (T[1, row-2] + T[2, row-2]) / 2 - 273.15   # at location 0.001 m, keep it
            ws_Sheet.cell(row, 3).value =T[0, row-2]-273.15 # temperature at surface
            ws_Sheet.cell(row, thermo_id).value = (T[int(Thermo_depth[thermo_id-4]*1000), row-2] + T[int(Thermo_depth[thermo_id-4]*1000 + 1), row-2]) / 2 - 273.15

         #ws_Sheet.cell(column, 3).value = (T[Thermo_depth[0]*1000, column-2] + T[Thermo_depth[0]*1000+1, column-2]) / 2 - 273.15  # Thermocouple 1
         #ws_Sheet.cell(column, 4).value = (T[Thermo_depth[1]*1000, column-2] + T[Thermo_depth[1]*1000+1, column-2]) / 2 - 273.15
         #ws_Sheet.cell(column, 5).value = (T[Thermo_depth[2]*1000, column-2] + T[Thermo_depth[2]*1000+1, column-2]) / 2 - 273.15


    wb_result.save(results_path + results_name)
    
    if (post_process == True):
        post_processing(Thermo_depth)
    if (Ucode == True):
        uCode(Thermo_depth)
        
    print("Simulation Complete")
   
def post_processing(Thermo_depth = []):
    results_path = shared.proj_folder_long

    results_name = "\\" + shared.proj_name + '-Simulation.xlsx'

    Test_data_path = shared.proj_folder_long
    Test_data_name = shared.proj_name + '-Raw.xlsx'
    Test_data = xl.load_workbook(Test_data_path + "\\" + Test_data_name)
    Data_sheet = Test_data['Sheet1']
    Max_Row_Data = Data_sheet.max_row
    #print(Max_Row_Data)

    #print(Data_sheet.cell(3, 2).value)
    #print(Data_sheet.cell(13, 3).value)

    Sim_data = xl.load_workbook(results_path + results_name)
    Sim_result = Sim_data['Sheet1']
    #date = str(Data_sheet.cell(4, 5).value)
    #time = str(Data_sheet.cell(4, 6).value)
    #print(date, time)

    Sim_First_Date = str(Sim_result.cell(3, 1).value)
    Sim_First_Time = str(Sim_result.cell(3, 2).value)
    Sim_Last_Date = str(Sim_result.cell(Sim_result.max_row, 1).value)
    Sim_Last_Time = str(Sim_result.cell(Sim_result.max_row, 2).value)

    Min_Year = int(Sim_First_Date[0])*1000 + int(Sim_First_Date[1])*100 + int(Sim_First_Date[2])*10 + int(Sim_First_Date[3])
    Max_Year = int(Sim_Last_Date[0])*1000 + int(Sim_Last_Date[1])*100 + int(Sim_Last_Date[2])*10 + int(Sim_Last_Date[3])# simulation year

    #Max_Month = int(Sim_Last_Date[5]) *10 + int(Sim_Last_Date[6])
    #Max_Day = int(Sim_Last_Date[8]) *10 + int(Sim_Last_Date[9])

    time_flag = 3
    Thermo_NO_flag = 1 # start from the first thermocouple number

    for time_row in range(3, Max_Row_Data+1):

        Date = str(Data_sheet.cell(time_row, 2).value)
        Year = int(Date[0])*1000 + int(Date[1])*100 + int(Date[2])*10 + int(Date[3])
        Month = int(Date[5]) *10 + int(Date[6])
        Day = int(Date[8]) *10 + int(Date[9])


        #Time = str(Data_sheet.cell(time_row, 3).value)
        #Hour = int(Time[0]) * 10 + int(Time[1])
        Hour = Data_sheet.cell(time_row, 3).value
        #AM_PM = (Time[9]) + (Time[10])
        '''
        if Hour == 24:
            Hour = 0

            if (Month == 1 or Month == 3 or Month == 5 or Month == 7 or Month == 8 or Month == 10) and Day == 31:
                Month = Month + 1
                Day = 1

            elif (Month == 4 or Month == 6 or Month == 9 or Month == 11) and Day == 30:
                Month = Month + 1
                Day = 1

            elif Month == 12 and Day == 31:
                Year = Year + 1
                Month = 1
                Day = 1

            elif ((Year%4==0) and (Year%100 !=0) or (Year%400==0)) and Month == 2 and Day == 29: # leap year

                Month = 3
                Day = 1

            elif (Year%4 !=0) and Month == 2 and Day == 28:
                Month = 3
                Day = 1

            else:
                Day = Day + 1
        '''
        print('Test = ', Year, Month, Day, Hour)
        #print('time_row=', time_row)

        #Thermo_NO = int(Data_sheet.cell(time_row, 4).value)
        #print('Thermo_NO=', Thermo_NO)

        # if Thermo_NO == Thermo_NO_flag, continue
        #if Thermo_NO != Thermo_NO_flag:
        #    time_flag = 3  # restart for next thermocouple
        #    Thermo_NO_flag = Thermo_NO  # assign Thermo_NO to Thermo_NO_flag

        if Min_Year <= Year <= Max_Year: # only iterate within the simulation year range

            for sim_time_row in range(time_flag, Sim_result.max_row+1):
                Sim_date = str(Sim_result.cell(sim_time_row, 1).value)

                Sim_Year = int(Sim_date[0])*1000 + int(Sim_date[1])*100 + int(Sim_date[2])*10 + int(Sim_date[3])
                Sim_Month = int(Sim_date[5]) *10 + int(Sim_date[6])
                Sim_Day = int(Sim_date[8]) *10 + int(Sim_date[9])

                Sim_Time = str(Sim_result.cell(sim_time_row, 2).value)

                Sim_Hour = int(Sim_Time[0])*10 + int(Sim_Time[1])
                Sim_AM_PM = (Sim_Time[9]) + (Sim_Time[10])


                if Sim_Hour == 12 and Sim_AM_PM == 'AM':  # 12 AM = 0
                    Sim_Hour = 0

                if Sim_AM_PM == 'PM' and Sim_Hour != 12:
                    Sim_Hour = Sim_Hour + 12


                if Sim_Year == Year and Sim_Month == Month and Sim_Day == Day and Sim_Hour == Hour:
                    date_time_str = str(Sim_date[0:10]) + ' ' + str(Sim_Time)
                    Sim_result.cell(sim_time_row-row_correction, 15).value = datetime.strptime(date_time_str, '%Y-%m-%d %I:%M:%S %p')
                    '''
                    if Sim_Hour < 10:
                        #date_time_str = str(Sim_Year) + '/' + str(Sim_Month) + '/' + str(Sim_Day) + ' ' + '0' + str(Sim_Hour) + ':00:00'
                        date_time_str = str(Sim_date[0:9]) + ' ' + str(Sim_Time)
                        Sim_result.cell(sim_time_row-row_correction, 15).value = datetime.strptime(date_time_str, '%Y-/%m/%dd %H:%M:%S')
                    #Date[0:10] + '-' + str(Hour) #+ str(Data_sheet.cell(time_row, 6).value[0:2]) #may need to shift it 1-2 hours ahead. Time zone?
                    if Sim_Hour >= 10:
                        date_time_str = str(Sim_Year) + '/' + str(Sim_Month) + '/' + str(Sim_Day) + ' ' + str(Sim_Hour) + ':00:00'
                        Sim_result.cell(sim_time_row-row_correction, 15).value = datetime.strptime(date_time_str, '%y/%m/%d %H:%M:%S')
                    '''
                    for Thermo_ID in range(0, len(Thermo_depth)):

                        Sim_result.cell(sim_time_row-row_correction, Thermo_ID+16).value = Data_sheet.cell(time_row, Thermo_ID+5).value  # [5:5+len(Thermo_depth)+1], assign test value to simulation file and may need to shift it 1-2 hours ahead. Time zone?

                        #time_flag = sim_time_row + 1

                        #print('time_flag=', time_flag, Thermo_NO_flag)
                        print('Sim=', Sim_Year, Sim_Month, Sim_Day, Sim_Hour)
                        #break

    Sim_data.save(results_path + results_name)

    # error calculation
    results_path = shared.proj_folder_long

    results_name = "\\" + shared.proj_name + '-Simulation.xlsx'

    Sim_data = xl.load_workbook(results_path + results_name)
    Sim_result = Sim_data['Sheet1']

    mean_abs_error = [0]*len(Thermo_depth)
    R2 = []
    a = [] # slope y = ax

    Thermo_NO_flag = len(Thermo_depth)

    for thermo_id in range(1, Thermo_NO_flag+1):

        x = [] # list for test value
        y = [] # list for simulation value
        date = []

        error_number = 0
        error_sum = 0

        for sim_time_row in range(3, Sim_result.max_row+1):
            if Sim_result.cell(sim_time_row, thermo_id + 15).value is not None:
                error_sum += abs(Sim_result.cell(sim_time_row, thermo_id + 15).value - Sim_result.cell(sim_time_row, thermo_id + 3).value)
                error_number += 1
                x.append(Sim_result.cell(sim_time_row, thermo_id + 15).value) # test value
                y.append(Sim_result.cell(sim_time_row, thermo_id + 3).value) # sim value
                date.append(Sim_result.cell(sim_time_row, 15).value) # date value
        print(error_number)
                #print(thermo_id, sim_time_row, type(Sim_result.cell(sim_time_row, thermo_id + 15).value))

        mean_abs_error[thermo_id-1] = round(error_sum / error_number, 5) # round to 5 decimals

        #slope, intercept, r_value, p_value, std_err = stats.linregress(x,y)
        #R2.append(r_value**2)

        x_reg = numpy.array(x).reshape(-1,1) # convert list to 2D array
        y_reg = numpy.array(y).reshape(-1,1)

        reg = LinearRegression(fit_intercept = False).fit(x_reg,y_reg) # fit_intercept = false (b = 0)
        R2.append(reg.score(x_reg, y_reg)) # coefficient of determination R2
        a.append(reg.coef_)


        # plot scatter points
        fig = plt.figure(num = 1) # figure size; figsize=(3,50)

        fig.add_subplot(1,Thermo_NO_flag,thermo_id) # total row, total column, position,xscale = "linear"

        plt.scatter(x,y,s=1) # s is circle size

        plt.xlabel('Test Temperature ($^\circ$C)', fontsize=10)
        plt.ylabel('Simulation Temperature ($^\circ$C)', fontsize=10)
        plt.title('Depth='+ str(Thermo_depth[thermo_id-1]) +'m') # add subplot title

        lower_lim = int(round(min(numpy.min(x_reg)-5,numpy.min(y_reg)-5)/5.0)*5.0) # obtain lower limit to the neareast 5
        upper_lim = int(round(max(numpy.max(x_reg)+5,numpy.max(y_reg)+5)/5.0)*5.0) # obtain upper limit to the neareast 5

        lim_range = [lower_lim,upper_lim] # define plot range integer

        plt.xlim(lim_range) # set x axis range, [-20,50]
        plt.ylim(lim_range)

        plt.plot(lim_range, lim_range, 'r') # add line of equity, x_lim, y_lim,color = 'r' , 'k'

        # plot regression
        #x_new = np.linspace(lower_lim,upper_lim,num=10) # create x for regression plot

        #y_new = LinearRegression().predict(x_new[:,np.newaxis])

        x_new = numpy.array(lim_range).reshape(-1,1)
        y_new = x_new * reg.coef_
        #print(x_new, y_new)

        plt.plot(x_new, y_new, '--k') # linear regression plot; black dashed line; reg.coef = a

        # also add equation, R2, and MAE

        equation = 'y=' + str(round(reg.coef_[0,0],3))+ 'x'  # reg.coef_ is array

        plt.text(lower_lim+10, upper_lim-5, equation, fontsize = 11)
        plt.text(lower_lim+10, upper_lim-7, '$R^{2}$=' + str(round(reg.score(x_reg, y_reg),3)), fontsize = 11) # round to 3 decimals
        plt.text(lower_lim+10, upper_lim-9, 'MAE=' + str(round(mean_abs_error[thermo_id-1],3))+'$^\circ$C', fontsize = 11)

        # plot the temperature with date
        fig2 = plt.figure()  #num = 2
        fig2.add_subplot() # if individual figures are preferred, do not define arguments; Thermo_NO_flag,1,thermo_id
        plt.plot(date, x, 'r', label='Test Temperature') # test temperature
        plt.plot(date, y,'b', label = 'Simulation Temperature') # simulation temperature
        plt.xlabel('Date', fontsize=12)
        plt.ylabel('Temperature ($^\circ$C)', fontsize=12)
        plt.title('Depth='+ str(Thermo_depth[thermo_id-1]) +'m') # add subplot title
        plt.ylim(lim_range)
        plt.legend(loc=2, prop={'size': 10})  # loc=0 is best, upper right = 1, upper left = 2

    plt.tight_layout()
    plt.show()

    print('MAE=', mean_abs_error)
    print('R2=', R2)
    print('Slope=',a)
    
def uCode(Thermo_depth = []):
    results_path = shared.proj_folder_long
    results_name = "\\" + shared.proj_name + '-Simulation.xlsx'
    Sim_data = xl.load_workbook(results_path + results_name)
    Sim_result = Sim_data['Sheet1']

    Ucode_path = shared.proj_folder_long
    Ucode_name = "\\" + shared.proj_name + '-Ucode.xlsx'

    try:
        Ucode_result = xl.load_workbook(Ucode_path + Ucode_name)
    except FileNotFoundError:
        Ucode_result = xl.Workbook()  # create wb as workbook
        Ucode_Sheet = Ucode_result.active  # define ws as worksheet
        Ucode_Sheet.title = 'Sheet1'
        Ucode_result.save(Ucode_path + Ucode_name)

    Ucode_result = xl.load_workbook(Ucode_path + Ucode_name)
    Ucode_Sheet = Ucode_result['Sheet1']

    Ucode_Sheet.cell(1, 1).value = 'Sim_results'
    Ucode_Sheet.cell(1, 3).value = 'Obsname'
    Ucode_Sheet.cell(1, 4).value = 'Test_results'
    Ucode_Sheet.cell(1, 5).value = 'Group Name'

    Ucode_row = 2   # start from row 2 as the first row is considered as headers in pd.DataFrame

    for thermo_id in range (1,len(Thermo_depth)+1):
        for time_row in range(3, Sim_result.max_row+1):
            if Sim_result.cell(time_row, thermo_id+15).value is not None:
                Ucode_Sheet.cell(Ucode_row, 1).value = Sim_result.cell(time_row, thermo_id+3).value
                Ucode_Sheet.cell(Ucode_row, 3).value = 'T'+str(thermo_id) + str(Sim_result.cell(time_row, 1).value)[0:4] + str(Sim_result.cell(time_row, 1).value)[5:7] + str(Sim_result.cell(time_row, 1).value)[8:10] + str(Sim_result.cell(time_row, 2).value)[0:2] + str(Sim_result.cell(time_row, 2).value)[9:11]
                Ucode_Sheet.cell(Ucode_row, 4).value = Sim_result.cell(time_row, thermo_id+15).value
                Ucode_Sheet.cell(Ucode_row, 5).value = 'T'+str(thermo_id)
                Ucode_row += 1
                print(thermo_id,time_row,Ucode_row,Sim_result.cell(time_row, thermo_id+3).value)

    Ucode_result.save(Ucode_path + Ucode_name)
    df = pd.read_excel(Ucode_path + Ucode_name)
    data = pd.DataFrame(df)  # first row contains headers
    Ucode_file = open(Ucode_path+"\\"+'sim_data.txt','w')
    #for data in df.columns[0]:
    Ucode_file.write(df[data.columns[0]].to_string(index=False)+'\n')
    Ucode_file.close()