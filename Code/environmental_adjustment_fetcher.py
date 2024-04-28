# -*- coding: utf-8 -*-
"""
Created on Sat Apr  6 17:50:13 2024

@author: drewm
"""

import openpyxl as xl
import shared

state = []
e1 = []
e6 = []

adjustment_file = xl.load_workbook(shared.proj_folder_long
                                   + "\\"
                                   + "Environmental Adjustments"
                                   + "\\"
                                   + "Adjustments (pure data).xlsx")

adjustment_data = adjustment_file['Sheet1']

for i in range(2, 30):
    state.append(str(adjustment_data.cell(i, 3).value) + " county, "
                 + str(adjustment_data.cell(i, 2).value))
    e1.append(adjustment_data.cell(i, 6).value)
    e6.append(adjustment_data.cell(i, 7).value)
