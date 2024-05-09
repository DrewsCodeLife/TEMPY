# -*- coding: utf-8 -*-
"""
Created on Sat Apr 27 15:58:54 2024

@author: drewm
"""

from collections import defaultdict
import openpyxl as xl
import shared
import math

# Each dictionary will contain 'year : year_average' pairs for the gui
max7day20mm = dict()
max3day20mm = dict()
max1day20mm = dict()

max7daySurf = dict()
max3daySurf = dict()
max1daySurf = dict()

min1daySurf = dict()

def run_calculations():
    resultBook = xl.load_workbook(shared.proj_folder_long
                                  + "\\"
                                  + shared.proj_name
                                  + '-Simulation.xlsx')

    resultData = resultBook["Sheet1"]

    # Starting sheetlength at -2 accounts for naming rows to be ignored
    sheetRowLength = -2
    sheetColLength = 0
    for cell in resultData['C']:
        if cell is None:
            break
        sheetRowLength += 1

    numDays = sheetRowLength / 24
    numYears = math.floor(numDays / 365)
    
    col20mm = 1
    for col in resultData.iter_cols():
        if col[0].value == '0.02 m':
            break
        col20mm += 1
    
    # Each of these dictionaries will contain:
    #   'year : list of n-day averages'
    max7day20mmVals = defaultdict(list)
    max3day20mmVals = defaultdict(list)
    max1day20mmVals = defaultdict(list)

    max7daySurfVals = defaultdict(list)
    max3daySurfVals = defaultdict(list)
    max1daySurfVals = defaultdict(list)

    min1dayVals = defaultdict(list)
    
    for year in range(0, numYears):
        # If the year is a leap year
        if year % 4 == 0 and (year % 100 != 0 or year % 400 == 0):
            for day in range(1, 367):
                values20mm = []
                valuesSurf = []
                for hour in range(2, 26):
                    values20mm.append(
                        resultData.cell(24 * (day - 1) + hour,
                                        col20mm).value
                    )
                    valuesSurf.append(
                        resultData.cell(24 * (day - 1) + hour, 3).value
                    )
                max1day20mmVals[year].append(max(values20mm))
                max1daySurfVals[year].append(max(valuesSurf))
                min1dayVals[year].append(min(values20mm))
                
            for i in range(0, 364):
                # stop calculating 7 day at 360
                max3day20mmVals[year].append(
                    sum(max1day20mmVals[year][i : i + 3]) / 3
                    )
                max3daySurfVals[year].append(
                    sum(max1daySurfVals[year][i : i + 3]) / 3
                    )
                
                if i < 361:
                    max7day20mmVals[year].append(
                        sum(max1day20mmVals[year][i : i + 7]) / 7
                        )
                    max7daySurfVals[year].append(
                        sum(max1daySurfVals[year][i : i + 7]) / 7
                        )
        else:
            for day in range(1, 366):
                values20mm = []
                valuesSurf = []
                for hour in range(2, 26):
                    values20mm.append(
                        resultData.cell(24 * (day - 1) + hour,
                                        col20mm).value
                    )
                    valuesSurf.append(
                        resultData.cell(24 * (day - 1) + hour, 3).value
                    )
                max1day20mmVals[year].append(max(values20mm))
                max1daySurfVals[year].append(max(valuesSurf))
                min1dayVals[year].append(min(values20mm))
                
            for i in range(0, 363):
                # stop calculating 7 day at 359
                max3day20mmVals[year].append(
                    sum(max1day20mmVals[year][i : i + 3]) / 3
                    )
                max3daySurfVals[year].append(
                    sum(max1daySurfVals[year][i : i + 3]) / 3
                    )
                
                if i < 360:
                    max7day20mmVals[year].append(
                        sum(max1day20mmVals[year][i : i + 7]) / 7
                        )
                    max7daySurfVals[year].append(
                        sum(max1daySurfVals[year][i : i + 7]) / 7
                        )
        max7day20mm[year] = max(max7day20mmVals[year]) 
        max3day20mm[year] = max(max3day20mmVals[year])
        max1day20mm[year] = max(max1day20mmVals[year])

        max7daySurf[year] = max(max7daySurfVals[year])
        max3daySurf[year] = max(max3daySurfVals[year])
        max1daySurf[year] = max(max1daySurfVals[year])

        min1daySurf[year] = max(min1dayVals[year])
    
    return numYears
